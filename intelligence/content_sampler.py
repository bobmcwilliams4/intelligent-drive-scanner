"""Content sampling and domain detection for files.

This module extracts just enough data from each file to intelligently classify it
without reading the entire file. Uses magic bytes, content samples, keywords, and
path analysis to auto-detect file domains and extract metadata.

Performance targets:
- Text files: <1ms
- Binary with magic: <5ms
- PDF/Office: ~10ms
- Batch mode: 100+ files/sec
"""

from __future__ import annotations

import asyncio
import hashlib
import re
import sqlite3
import time
from collections import Counter
from pathlib import Path
from typing import Any

from loguru import logger

from config import (
    BINARY_EXTENSIONS,
    CONTENT_DOMAIN_RULES,
    EXTENSION_DOMAIN,
    EXTENSION_MIME,
    FILE_SIGNATURES,
    KEYWORD_LIMIT,
    PATH_DOMAIN_RULES,
    SAMPLE_SIZE,
    SIGNATURE_SIZE,
    TEXT_EXTENSIONS,
)
from storage.models import FileSample

# Try to import xxhash, fallback to hashlib
try:
    import xxhash
    HAS_XXHASH = True
except ImportError:
    HAS_XXHASH = False
    logger.warning("xxhash not installed, using md5 for fast hashing")

# Optional dependencies for rich content extraction
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Common English stopwords to filter out
STOPWORDS = {
    "the", "be", "to", "of", "and", "a", "in", "that", "have", "i",
    "it", "for", "not", "on", "with", "he", "as", "you", "do", "at",
    "this", "but", "his", "by", "from", "they", "we", "say", "her", "she",
    "or", "an", "will", "my", "one", "all", "would", "there", "their",
    "what", "so", "up", "out", "if", "about", "who", "get", "which", "go",
    "me", "when", "make", "can", "like", "time", "no", "just", "him", "know",
    "take", "people", "into", "year", "your", "good", "some", "could", "them",
    "see", "other", "than", "then", "now", "look", "only", "come", "its", "over",
}


class ContentSampler:
    """Samples file content and extracts metadata for intelligent classification."""

    def __init__(self) -> None:
        """Initialize content sampler."""
        self._sample_cache: dict[Path, FileSample] = {}
        self._hash_threshold = 1_073_741_824  # 1GB - skip hashing above this

    async def sample_file(self, path: Path) -> FileSample:
        """Sample a file and extract metadata.

        Args:
            path: Path to file to sample

        Returns:
            FileSample with metadata, signature, content sample, keywords, domain
        """
        start_time = time.perf_counter()

        try:
            # Basic file stats
            stat = path.stat()
            size = stat.st_size
            modified = stat.st_mtime
            extension = path.suffix.lower()

            # Read file signature (first 16 bytes)
            signature = self._read_signature(path)

            # Detect MIME type from signature + extension
            mime_type = self._detect_mime(path, bytes.fromhex(signature))

            # Determine if binary based on extension and signature
            is_binary = extension in BINARY_EXTENSIONS or not self._is_text_mime(mime_type)

            # Read content sample and extract keywords (skip for large binaries)
            content_sample = None
            keywords: list[str] = []
            if size < 100_000_000:  # Only sample files < 100MB
                content_sample, keywords = self._read_content_sample(path, mime_type)

            # Auto-detect domain
            domain, confidence = self._detect_domain(path, extension, content_sample, keywords)

            # Compute hashes (skip for very large files)
            sha256_hash = ""
            fast_hash = ""
            if size < self._hash_threshold:
                sha256_hash, fast_hash = self._compute_hashes(path)

            # Calculate sampling time
            sample_time_ms = (time.perf_counter() - start_time) * 1000

            return FileSample(
                path=str(path),
                size=size,
                extension=extension,
                mime_type=mime_type,
                signature=signature,
                is_binary=is_binary,
                content_sample=content_sample,
                keywords=keywords,
                domain=domain,
                domain_confidence=confidence,
                sha256=sha256_hash,
                fast_hash=fast_hash,
                modified_time=modified,
                sample_time_ms=sample_time_ms,
            )

        except PermissionError:
            logger.warning(f"Permission denied: {path}")
            return self._create_error_sample(path, "Permission denied")
        except OSError as e:
            logger.warning(f"OS error sampling {path}: {e}")
            return self._create_error_sample(path, str(e))
        except Exception as e:
            logger.error(f"Unexpected error sampling {path}: {e}")
            return self._create_error_sample(path, str(e))

    def _read_signature(self, path: Path) -> str:
        """Read file signature (first 16 bytes as hex).

        Args:
            path: Path to file

        Returns:
            Hex string of first 16 bytes (or less if file is smaller)
        """
        try:
            with path.open("rb") as f:
                signature_bytes = f.read(SIGNATURE_SIZE)
                return signature_bytes.hex()
        except Exception as e:
            logger.debug(f"Could not read signature from {path}: {e}")
            return ""

    def _detect_mime(self, path: Path, signature: bytes) -> str:
        """Detect MIME type from magic bytes and extension.

        Args:
            path: Path to file
            signature: First 16 bytes of file

        Returns:
            MIME type string
        """
        # Check magic bytes first
        for magic_hex, mime in FILE_SIGNATURES.items():
            magic_bytes = bytes.fromhex(magic_hex)
            if signature.startswith(magic_bytes):
                return mime

        # Fallback to extension-based detection
        extension = path.suffix.lower()
        return EXTENSION_MIME.get(extension, "application/octet-stream")

    def _is_text_mime(self, mime: str) -> bool:
        """Check if MIME type indicates text content."""
        return mime.startswith("text/") or mime in {
            "application/json",
            "application/xml",
            "application/javascript",
            "application/x-python",
        }

    def _read_content_sample(self, path: Path, mime: str) -> tuple[str | None, list[str]]:
        """Read content sample and extract keywords.

        Args:
            path: Path to file
            mime: MIME type

        Returns:
            Tuple of (content_sample, keywords)
        """
        extension = path.suffix.lower()

        # Handle specific file types with specialized extractors
        if extension == ".pdf":
            return self._sample_pdf(path)
        elif extension == ".docx":
            return self._sample_docx(path)
        elif extension == ".xlsx":
            return self._sample_xlsx(path)
        elif extension in {".db", ".sqlite", ".sqlite3"}:
            return self._sample_sqlite(path)
        elif extension in {".py", ".js", ".ts", ".go", ".rs", ".java", ".cpp", ".c", ".h"}:
            return self._sample_code(path)
        elif extension == ".json":
            return self._sample_json(path)
        elif extension in TEXT_EXTENSIONS or self._is_text_mime(mime):
            return self._sample_text(path)
        else:
            # Binary file - no content sample
            return None, []

    def _sample_text(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample plain text file."""
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(SAMPLE_SIZE)
                keywords = self._extract_keywords_from_text(sample)
                return sample, keywords
        except Exception as e:
            logger.debug(f"Could not sample text from {path}: {e}")
            return None, []

    def _sample_code(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample code file with specialized keyword extraction."""
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(SAMPLE_SIZE)
                keywords = self._extract_keywords_from_code(sample)
                return sample, keywords
        except Exception as e:
            logger.debug(f"Could not sample code from {path}: {e}")
            return None, []

    def _sample_json(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample JSON file and extract keys."""
        try:
            import json
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                content = f.read(SAMPLE_SIZE)
                # Try to parse as JSON to extract keys
                try:
                    data = json.loads(content)
                    keywords = self._extract_json_keys(data)
                except json.JSONDecodeError:
                    # Partial JSON, just extract text keywords
                    keywords = self._extract_keywords_from_text(content)
                return content, keywords
        except Exception as e:
            logger.debug(f"Could not sample JSON from {path}: {e}")
            return None, []

    def _sample_pdf(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample PDF file (first 2 pages)."""
        if not HAS_FITZ:
            return None, []

        try:
            doc = fitz.open(path)
            text_parts = []
            # Extract text from first 2 pages
            for page_num in range(min(2, len(doc))):
                page = doc[page_num]
                text_parts.append(page.get_text())
            doc.close()

            sample = " ".join(text_parts)[:SAMPLE_SIZE]
            keywords = self._extract_keywords_from_text(sample)
            return sample, keywords
        except Exception as e:
            logger.debug(f"Could not sample PDF {path}: {e}")
            return None, []

    def _sample_docx(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample DOCX file (first 20 paragraphs)."""
        if not HAS_DOCX:
            return None, []

        try:
            doc = Document(path)
            paragraphs = []
            for para in doc.paragraphs[:20]:
                if para.text.strip():
                    paragraphs.append(para.text)

            sample = " ".join(paragraphs)[:SAMPLE_SIZE]
            keywords = self._extract_keywords_from_text(sample)
            return sample, keywords
        except Exception as e:
            logger.debug(f"Could not sample DOCX {path}: {e}")
            return None, []

    def _sample_xlsx(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample XLSX file (first sheet headers + 10 rows)."""
        if not HAS_OPENPYXL:
            return None, []

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 11:  # Header + 10 data rows
                    break
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                rows.append(row_text)
            wb.close()

            sample = " ".join(rows)[:SAMPLE_SIZE]
            keywords = self._extract_keywords_from_text(sample)
            return sample, keywords
        except Exception as e:
            logger.debug(f"Could not sample XLSX {path}: {e}")
            return None, []

    def _sample_sqlite(self, path: Path) -> tuple[str | None, list[str]]:
        """Sample SQLite database (table and column names)."""
        try:
            conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            cursor = conn.cursor()

            # Get table names
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]

            # Get column names from first few tables
            keywords = []
            for table in tables[:5]:
                cursor.execute(f"PRAGMA table_info({table})")
                columns = [row[1] for row in cursor.fetchall()]
                keywords.extend(columns)

            keywords.extend(tables)
            conn.close()

            sample = f"Tables: {', '.join(tables[:10])}"
            return sample, keywords[:KEYWORD_LIMIT]
        except Exception as e:
            logger.debug(f"Could not sample SQLite {path}: {e}")
            return None, []

    def _extract_keywords_from_text(self, text: str) -> list[str]:
        """Extract top keywords from text content.

        Args:
            text: Text content

        Returns:
            List of top keywords by frequency
        """
        # Remove punctuation, split on whitespace
        words = re.findall(r'\b[a-zA-Z_][a-zA-Z0-9_]*\b', text.lower())

        # Filter stopwords and short words
        filtered = [w for w in words if w not in STOPWORDS and len(w) >= 3]

        # Count frequency
        counter = Counter(filtered)

        # Return top N
        return [word for word, _ in counter.most_common(KEYWORD_LIMIT)]

    def _extract_keywords_from_code(self, text: str) -> list[str]:
        """Extract keywords from code (imports, functions, classes).

        Args:
            text: Code content

        Returns:
            List of keywords (imports, function names, class names)
        """
        keywords = []

        # Extract import statements (Python)
        imports = re.findall(r'^\s*(?:from|import)\s+([\w.]+)', text, re.MULTILINE)
        keywords.extend(imports)

        # Extract function definitions
        functions = re.findall(r'^\s*(?:def|function|func|fn)\s+(\w+)', text, re.MULTILINE)
        keywords.extend(functions)

        # Extract class definitions
        classes = re.findall(r'^\s*class\s+(\w+)', text, re.MULTILINE)
        keywords.extend(classes)

        # Extract common code keywords
        code_keywords = re.findall(r'\b(?:async|await|return|yield|const|let|var|struct|interface|type|enum)\b', text)
        keywords.extend(code_keywords)

        # Add general text keywords as fallback
        text_keywords = self._extract_keywords_from_text(text)
        keywords.extend(text_keywords[:20])

        return keywords[:KEYWORD_LIMIT]

    def _extract_json_keys(self, data: Any, depth: int = 0) -> list[str]:
        """Recursively extract keys from JSON data.

        Args:
            data: Parsed JSON data
            depth: Current recursion depth

        Returns:
            List of keys
        """
        if depth > 3:  # Limit recursion depth
            return []

        keys = []
        if isinstance(data, dict):
            keys.extend(data.keys())
            for value in data.values():
                keys.extend(self._extract_json_keys(value, depth + 1))
        elif isinstance(data, list):
            for item in data[:5]:  # Sample first 5 items
                keys.extend(self._extract_json_keys(item, depth + 1))

        return keys[:KEYWORD_LIMIT]

    def _detect_domain(
        self,
        path: Path,
        extension: str,
        content: str | None,
        keywords: list[str],
    ) -> tuple[str, float]:
        """Auto-detect file domain using multi-stage analysis.

        Priority:
        1. Extension-based (instant, high confidence)
        2. Path-based (regex on directories)
        3. Content-based (keyword matching)
        4. Default to UNKNOWN

        Args:
            path: Path to file
            extension: File extension
            content: Content sample (if available)
            keywords: Extracted keywords

        Returns:
            Tuple of (domain, confidence)
        """
        # Stage 1: Extension-based detection (highest confidence)
        if extension in EXTENSION_DOMAIN:
            return EXTENSION_DOMAIN[extension], 0.95

        # Stage 2: Path-based detection
        path_str = str(path).lower()
        for pattern, domain in PATH_DOMAIN_RULES:
            if re.search(pattern, path_str):
                return domain, 0.85

        # Stage 3: Content-based detection
        if content or keywords:
            keyword_str = " ".join(keywords).lower()
            content_str = (content or "").lower()
            combined = f"{keyword_str} {content_str}"

            for pattern, domain in CONTENT_DOMAIN_RULES:
                if re.search(pattern, combined):
                    return domain, 0.70

        # Stage 4: Default
        return "UNKNOWN", 0.0

    def _compute_hashes(self, path: Path) -> tuple[str, str]:
        """Compute SHA-256 and fast hash (xxHash or MD5).

        Args:
            path: Path to file

        Returns:
            Tuple of (sha256_hex, fast_hash_hex)
        """
        try:
            sha256 = hashlib.sha256()
            if HAS_XXHASH:
                fast = xxhash.xxh64()
            else:
                fast = hashlib.md5()

            with path.open("rb") as f:
                while chunk := f.read(8192):
                    sha256.update(chunk)
                    fast.update(chunk)

            return sha256.hexdigest(), fast.hexdigest()
        except Exception as e:
            logger.debug(f"Could not compute hashes for {path}: {e}")
            return "", ""

    def _create_error_sample(self, path: Path, error_msg: str) -> FileSample:
        """Create a minimal FileSample for files that errored during sampling.

        Args:
            path: Path to file
            error_msg: Error message

        Returns:
            FileSample with minimal data
        """
        try:
            stat = path.stat()
            size = stat.st_size
            modified = stat.st_mtime
        except Exception:
            size = 0
            modified = 0.0

        return FileSample(
            path=str(path),
            size=size,
            extension=path.suffix.lower(),
            mime_type="application/octet-stream",
            signature="",
            is_binary=True,
            content_sample=None,
            keywords=[],
            domain="UNKNOWN",
            domain_confidence=0.0,
            sha256="",
            fast_hash="",
            modified_time=modified,
            sample_time_ms=0.0,
        )

    async def sample_batch(self, paths: list[Path]) -> list[FileSample]:
        """Sample multiple files in parallel.

        Args:
            paths: List of file paths to sample

        Returns:
            List of FileSamples
        """
        # Use semaphore to limit concurrency
        semaphore = asyncio.Semaphore(100)

        async def sample_with_semaphore(path: Path) -> FileSample:
            async with semaphore:
                return await self.sample_file(path)

        tasks = [sample_with_semaphore(path) for path in paths]
        return await asyncio.gather(*tasks)
